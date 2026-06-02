import os
import io
from flask import Flask, request, jsonify
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

app = Flask(__name__)

def get_drive_service(readonly=True):
    scope = 'https://www.googleapis.com/auth/drive.readonly' if readonly else 'https://www.googleapis.com/auth/drive'
    creds = Credentials(
        token=None,
        refresh_token=os.environ['GOOGLE_REFRESH_TOKEN'],
        client_id=os.environ['GOOGLE_CLIENT_ID'],
        client_secret=os.environ['GOOGLE_CLIENT_SECRET'],
        token_uri='https://oauth2.googleapis.com/token',
        scopes=[scope]
    )
    creds.refresh(Request())
    return build('drive', 'v3', credentials=creds)

def export_file_content(file_id, mime_type):
    service = get_drive_service(readonly=True)
    try:
        # --- Archivos nativos de Google ---
        if mime_type == 'application/vnd.google-apps.document':
            result = service.files().export(fileId=file_id, mimeType='text/plain').execute()
            return result.decode('utf-8')[:4000]

        elif mime_type == 'application/vnd.google-apps.spreadsheet':
            result = service.files().export(fileId=file_id, mimeType='text/csv').execute()
            return result.decode('utf-8')[:4000]

        elif mime_type == 'application/vnd.google-apps.presentation':
            result = service.files().export(fileId=file_id, mimeType='text/plain').execute()
            return result.decode('utf-8')[:4000]

        # --- Archivos binarios: Excel, Word, PDF ---
        else:
            request_obj = service.files().get_media(fileId=file_id)
            buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(buffer, request_obj)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            buffer.seek(0)
            raw_bytes = buffer.read()

            # Excel (.xlsx, .xls)
            if mime_type in [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel'
            ]:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    lines.append(f"=== Hoja: {sheet} ===")
                    for row in ws.iter_rows(values_only=True):
                        row_str = ', '.join([str(c) if c is not None else '' for c in row])
                        if row_str.strip(','):
                            lines.append(row_str)
                return '\n'.join(lines)[:4000]

            # Word (.docx)
            elif mime_type in [
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/msword'
            ]:
                import docx
                doc = docx.Document(io.BytesIO(raw_bytes))
                text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
                return text[:4000]

            # PDF
            elif mime_type == 'application/pdf':
                import pdfplumber
                text_parts = []
                with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
                    for page in pdf.pages:
                        text_parts.append(page.extract_text() or '')
                return '\n'.join(text_parts)[:4000]

            # Texto plano, CSV, etc.
            elif mime_type.startswith('text/'):
                return raw_bytes.decode('utf-8', errors='ignore')[:4000]

            else:
                return f"Tipo de archivo no soportado: {mime_type}"

    except Exception as e:
        return f"Error extrayendo contenido: {str(e)}"

@app.route('/extract', methods=['GET'])
def extract():
    file_id = request.args.get('file_id', '')
    mime_type = request.args.get('mime_type', '')
    if not file_id:
        return jsonify({'error': 'file_id requerido'}), 400
    content = export_file_content(file_id, mime_type)
    return jsonify({'content': content})

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
